import pyautogui, time, random
import openpyxl, json, sys, re
import pyinputplus as pyip
from datetime import datetime

from pathlib import Path
import json


email_subjects = [
    "Join Reobrix: Elevate the World of Building Blocks as Our Next MOC Designer!",
    "Reobrix is Searching for Creative Minds in MOC Design – Let’s Collaborate!",
    "Got a Passion for Building? Reobrix Wants You as Our MOC Designer!",
    "Enhance Your Creative Journey with Reobrix – Seeking MOC Designers Now!",
    "Collaborate with Reobrix to Shape the Future of Building Blocks!",
    "Reobrix Needs Your Expertise in MOC Design – Let's Build Something Great!",
    "Calling All Building Block Innovators – Partner with Reobrix as a MOC Designer!",
    "Are You the Next Influencer for Reobrix? Join Us in Redefining Building Blocks!",
    "Help Us Inspire Builders Everywhere – Become a Reobrix MOC Designer!",
    "Your Creativity Is Wanted: Join Reobrix’s Team of Elite MOC Designers!"
]

# Print the list to verify its content
print(email_subjects)


def extractDataFromExcel(templatefile):
    try:
        wb = openpyxl.load_workbook(templatefile)
        sheet = wb.active

        cache_list =[]
        for i in range(2, sheet.max_row+1): # Go through every  row:
            temp_list_for_row = []
            for j in range(1, sheet.max_column+1):
                temp_list_for_row.append(sheet.cell(row=i, column=j).value)

            cache_list.append(temp_list_for_row)
        return cache_list
    except:
        print('Make sure the excel template exists.')
        sys.exit()    


def get_data_from_path(path):
    if path.exists():
        contents = path.read_text()
        extracted_ds = json.loads(contents)
        return extracted_ds
    else:
        return None 
    
def save_new_data_2path(data, path):
    data_str = json.dumps(data)
    path.write_text(data_str)
    # print(f'{data_str} saved.')


def string_to_xy_tuples(input_string: str):
    # take a str input like "553,223\n553,225\n" into [(553, 223), (553, 225)]
    lines = input_string.strip().split('\n')
    result = []
    for line in lines:
        x, y = map(int, line.split(','))
        result.append((x, y))
    return result





def gcd(m, n):
    while m % n != 0:
        m, n = n, m % n
    return n

class Fraction:
    def __init__(self, top, bottom):
        self.num = top
        self.den = bottom

    def __str__(self):
        return "{:d}/{:d}".format(self.num, self.den)

    def __eq__(self, other_fraction):
        first_num = self.num * other_fraction.den
        second_num = other_fraction.num * self.den

        return first_num == second_num

    def __add__(self, other_fraction):
        new_num = self.num * other_fraction.den \
        + self.den * other_fraction.num
        new_den = self.den * other_fraction.den
        cmmn = gcd(new_num, new_den)
        return Fraction(new_num // cmmn, new_den // cmmn)

    def show(self):
        print("{:d}/{:d}".format(self.num, self.den))



class RoundFunction:
    def __init__(self,n, f1, f2):
        self.n = n
        self.counter = 0
        
    def __iter__(self):
        return self
    
    def __next__(self):
        if self.counter < self.n-1:
            f1()
            self.counter += 1
        else:
            for i in range(self.n-1):
                f2()
                self.counter -= 1


def f1():
    pyautogui.hotkey("shift", "down")
    print('Pressed: shift-down')

def f2():
    pyautogui.hotkey("shift", "up")
    print('Pressed: shift-up')

def foxmail(n=3):
    '''import pyautogui; pyautogui.mouseInfo()'''

    import json, random

    with open('my_list.json', 'r') as file:
        mylist = json.load(file)
    email_number = n
    cache_list = extractDataFromExcel(r'temp_email\\gui_temp.xlsx')
    history_json = r'temp_email\\history_json.json'
    path = Path(history_json)

    # email_number = 3
    r = RoundFunction(email_number, f1, f2)

    print('Extracted info done.')
    print("Rows: ", len(cache_list))

    print(cache_list)

    '''136,40
95,85
93,153
15,247'''

    # p1 = 180,47
    # p2 = 118,104
    # p3 = 427,184
    # p4 = 21,312
    p1, p2, p3, p4 = [(136, 40), (95, 85), (93, 153), (15, 247)]
    def autowork(email, words, title):

        if not email:
            email = ''
        if not words:
            words = ''
        if not title:
            title = ''
        time.sleep(3)
        pyautogui.click(p1) # send key
        time.sleep(0.9)

        pyautogui.click(p2) # receivers
        pyautogui.write(email) # writing emails address
        time.sleep(0.9)
        # pyautogui.click(102, 154) # subject
        # pyautogui.write(title)# x biaoti
        # time.sleep(0.9)

        time.sleep(1.2)
        pyautogui.click(p3) # titile
        # pyautogui.write(title)
        pyautogui.write(random.choice(email_subjects))

        time.sleep(2)
        pyautogui.click(p4) # content
        

        # pyautogui.write(words)
        pyautogui.write(random.choice(mylist))


        # pyautogui.click(31, 42) #send
        pyautogui.hotkey('ctrl', 'enter')
    import pyinputplus as pyip
    max_time = pyip.inputInt(prompt="Enter your total time default 180secs: ", min=20, max=600, blank=True)
    if not max_time:
        max_time = 180
    print(f'max random time to wait is: {max_time}')  
    time.sleep(3)
    # special_count = 0
    for email_tuple in cache_list[:]:
        # special_count += 1
        emails_sent = get_data_from_path(path)
        if emails_sent == None:
            emails_sent = []
        email = email_tuple[0]
        name = email_tuple[1]
        title = email_tuple[2]
        words = email_tuple[3]

        if email in emails_sent:
            print(f'{email} already done.')
            continue
        else:
            emails_sent.append(email)
            # print(emails_sent)
            save_new_data_2path(emails_sent, path)

            autowork(email, words, title)

            sleep_time = random.randint(30, max_time)
            time.sleep(sleep_time//email_number)

            # pyautogui.click((58,17)); time.sleep(1)
            if r.n > 1:
                next(r)

import json


if __name__ == "__main__":
    number = input('How many accounts: ')
    if not number:
        number = 3
    else:
        number = int(number)
    foxmail(number)
    # outlook_web()
    pass