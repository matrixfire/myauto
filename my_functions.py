import openpyxl, re, time, random, sys, json, os, datetime
import pyautogui, openpyxl, threading
import pyinputplus as pyip
from pathlib import Path
from openpyxl.utils import get_column_letter, column_index_from_string
import pyperclip

import requests
from bs4 import BeautifulSoup

from openpyxl.styles import Font

from ordered_set import OrderedSet






import re
import pyperclip
from ordered_set import OrderedSet


emailRegex = re.compile(r'''(
    [a-zA-Z0-9._%+-]+   # username
    @ 
    [a-zA-Z0-9.-]+  # domain name
    (\.[a-zA-Z]{2,4}) 
    )''', re.VERBOSE)


def find_emails(txt=''):
    if not txt:
        txt = pyperclip.paste()
    # lt = [i[0] for i in emailRegex.findall(txt)]
    no_dup = OrderedSet()
    for i in emailRegex.findall(txt):
        emailAdd = i[0].lower()
        if emailAdd not in no_dup:
            print(emailAdd)
            no_dup.add(emailAdd)
    result_list = list(no_dup)
    print(f'\n{len(result_list)} emails copied to clipboard.\n')
    pyperclip.copy('\n'.join(result_list))
    # print(f'\nTotal: {len(no_dup)}')
    return result_list














def about_match_object_group():
    ''' 
    practices for search method of Regex objects
    the following symbols need to escape
    .  ^  $  *  +  ?  {  }  [  ]  \  |  (  )

    '''
    test_text = 'My number is 415-555-4242.'
    phoneNumRegex1 = re.compile(r'(\d\d\d)-(\d\d\d-\d\d\d\d)')
    phoneNumRegex2 = re.compile(r'\d\d\d-\d\d\d-\d\d\d\d')
    phoneNumRegex3 = re.compile(r'((\d\d\d)-(\d\d\d-\d\d\d\d))')
    mo = phoneNumRegex1.search(test_text)
    mo2 = phoneNumRegex2.search(test_text)
    mo3 = phoneNumRegex3.search(test_text)

    print(mo.group(), mo.group(0), mo.group(1), mo.group(2), sep=',')
    print(mo2.group(), mo2.group(0), sep=',')
    print(mo3.group(), mo3.group(0), mo3.group(1), mo3.group(2), mo3.group(3), sep=',')
    print(mo3.groups())

# about_match_object_group()



def about_findall_method():
    '''
    When called on a regex with no groups, such as \d\d\d-\d\d\d-\d\d\d\d, 
    the method findall() returns a list of string matches, such as ['415-555-9999', '212-555-0000'].
    When called on a regex that has groups, such as (\d\d\d)-(\d\d\d)-(\d\d\d\d), 
    the method findall() returns a list of tuples of strings (one string for each group), 
    such as [('415', '555', '9999'), ('212', '555', '0000')].
    '''
    test_text = 'Cell: 415-555-9999 Work: 212-555-0000'
    phoneNumRegex1 = re.compile(r'\d\d\d-\d\d\d-\d\d\d\d') # has no groups
    phoneNumRegex2 = re.compile(r'((\d\d\d)-(\d\d\d)-(\d\d\d\d))') # has no groups
    lt1 = phoneNumRegex1.findall(test_text)
    lt2 = phoneNumRegex2.findall(test_text)
    print(lt1)
    print(lt2)

'''
\s: Any space, tab, or newline character. (Think of this as matching “space” characters.)
\d: Any numeric digit from 0 to 9.
\w: Any letter, numeric digit, or the underscore character. (Think of this as matching “word” characters.)

re.DOTALL
re.IGNORECASE (or re.I)
re.VERBOSE

re.VERBOSE | re.DOTALL | re.I

'''

def about_sub_method():

    test_text = 'Agent Amy and Agent Alice told Agent Carol that Agent Eve knew Agent Bob was a double agent, and he sold out agent Christ.'
    # make the above to A**** told C**** that E**** knew B**** was a double agent.', with first char and * numbers the same

    namesRegex = re.compile(r'Agent (\w)(\w*)', re.I)

    lt = namesRegex.findall(test_text)
    lt = [''.join(i) for i in lt]

    d = {i: '*' * (len(i)-1) for i in lt}
    sub_text = r'\1{\1\2}'
    new_text = namesRegex.sub(sub_text, test_text)
    print(new_text)
    print(new_text.format(**d))



'''







copied_text = str(pyperclip.paste())
# copied_text = 'a32221@qq.com'
lt = [i[0] for i in emailRegex.findall(copied_text)]
# print(lt)
pyperclip.copy('\n'.join(lt))
print(str(pyperclip.paste()))
'''

easyPhoneReg = re.compile(r'\+[0-9 ]+\n')

urlReg = re.compile('https?:\\/\\/(?:www\\.)[-a-zA-Z0-9@:%._\\+~#=]{1,256}\\.[a-zA-Z0-9()]{1,6}\\b(?:[-a-zA-Z0-9()@:%_\\+.~#?&\\/=]*)')

def storeList2Excel(alist, filename): # alist is like [['a', 'b''], ['c', 'd']]

    someFont = Font(italic=True, bold=True)
    wb = openpyxl.Workbook()
    print(alist[:2]) # test
    sheet = wb.active

    sheet.freeze_panes = "B2"

    # print('Rows: ', len(alist))
    headerList = ["AAA", "BBB"]
    headerList = list(map(lambda a: a.title(), headerList))

    for column_index, cell_data in enumerate(headerList):
        # print(cell_data)
        sheet.cell(row=1, column=column_index+1).value = cell_data 
        sheet.cell(row=1, column=column_index+1).font = someFont 

    # below is real work    
    for row_index, rowList in enumerate(alist):
        # print("columns", len(rowList))
        for column_index, cell_data in enumerate(rowList):
            # print(cell_data)
            sheet.cell(row=row_index+2, column=column_index+1).value = cell_data

    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                max_length =  max((dims.get(cell.column, 0), len(str(cell.value)), 5)) 
                dims[cell.column_letter] = min(max_length, 50)
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value


    wb.save(filename)
    return filename




def print_list(alist):
    for i in alist:
        print(i.strip())
    print(f'\nTotal: {len(alist)}')
    pyperclip.copy('\n'.join(alist))

def get_text(filename):
    with open(filename, encoding='utf-8') as fo:
        text = fo.read()
    return text

def get_text_clipboard():
    return pyperclip.paste()


au_pRg = re.compile(r'(\+61 [45]\d\d \d\d\d \d\d\d)')



def get_list_fr_file(filename):
    with open(filename) as fo:
        temp_list = fo.readlines()
        result = list(filter(lambda a: len(a)>0, map(lambda a: a.strip().lower(), temp_list)))
    return result





def get_words_frequency(txt, no_words=''):
    txt = txt.lower()

    no_list = []
    if  no_words:
        no_list = get_list_fr_file(no_words)

    for ch in '!"#$%&()*+,-./:;<=>?@[\\]^_‘{|}~0123456789–©':
        txt = txt.replace(ch, ' ')

    word_list = txt.split()
    word_list = list(filter(lambda x: len(x) > 1, word_list))

    cache_dict = {}
    # print_list(word_list)
    for word in word_list:
        if word in no_list:
            continue
        cache_dict[word] = cache_dict.get(word, 0) + 1
    cache_list = list(cache_dict.items())
    cache_list.sort(key=lambda x: x[1], reverse=True)
    txt_copy = ''
    for word, num in cache_list:
        print(f'{word}, {num}')
        txt_copy += (word+'\n')

    pyperclip.copy(txt_copy)



def find_phone_numbers(pRg, txt):

    no_dup = set()
    for i in pRg.findall(txt):
        phoneNumber = i.lower()
        if phoneNumber not in no_dup:
            print(phoneNumber)
            no_dup.add(phoneNumber)
    
    result_list = list(no_dup)
    pyperclip.copy('\n'.join(result_list))
    print(f'\n{len(result_list)} phone numbers copied to clipboard.\n')
    # print(f'\nTotal: {len(no_dup)}')
    return result_list    




def efilter(email_str):
    result = 0
    username = email_str.split('@')[0]
    domainname = email_str.split('@')[1]
    if len(username) > 20 or len(domainname) > 30:
        result = 1

    if '-' in email_str:
        result = 1
    

    counter = 0
    for i in list('0123456789')+['+', '-']:
        if i in email_str:
            counter += 1
            # print('ii')
    # print(counter)
    if counter > 3:
        result = 1
    if counter/len(username) > 0.3:
        result = 1     

    if 'postmaster' in email_str or 'spam' in email_str:
        result = 1

    if domainname.count('.') >2:
        result = 1


    if result == 1:
        print(f'BAD EMAIL {email_str}')
        return True
    else:
        return False

# print(efilter('1661332010.13546@mail.freightlinks.net'))



def find_emails(txt, display=True, efilter=None):
    # lt = [i[0] for i in emailRegex.findall(txt)]
    no_dup = OrderedSet()
    # no_dup = set()
    for i in emailRegex.findall(txt):
        emailAdd = i[0].lower()
        if efilter is None:
            pass
        else:
            if efilter(emailAdd):
                continue

        if emailAdd not in no_dup:
            if display:
                print(emailAdd)
            no_dup.add(emailAdd)
    
    result_list = list(no_dup)
    
    if display:
        print(f'\n{len(result_list)} emails copied to clipboard.\n')
        pyperclip.copy('\n'.join(result_list))

    # print(f'\nTotal: {len(no_dup)}')
    return result_list




def extractDataFromExcel(templatefile):
    try:
        wb = openpyxl.load_workbook(templatefile)
        sheet = wb.active

        cache_list =[]
        for i in range(2, sheet.max_row+1): # Go through every  row:
            temp_list_for_row = []
            for j in range(1, sheet.max_column+1):
                temp_list_for_row.append(sheet.cell(row=i, column=j).value)

            cache_list.append(temp_list_for_row)
        return cache_list
    except:
        print('Make sure the excel template exists.')
        sys.exit()    





def get_emails_from_files(temp_dir):
    

    file_failed = []

    os.chdir(temp_dir)
    email_set = set()
    for filename in os.listdir('.'):
        if not (filename.endswith('.eml') or filename.endswith('.txt')):
            continue
        try:
            with open(filename, 'r', encoding='ISO-8859-1') as f_obj:
                temp_emails_found = find_emails(f_obj.read())
                for e in temp_emails_found:
                    email_set.add(e)
        except:
            file_failed.append(filename)

    pyperclip.copy('\n'.join(list(email_set)))
    with open('resultx.txt', 'w') as f:
        f.write('\n'.join(list(email_set)))

    
    print(f'\n{len(email_set)} emails copied to clipboard.\n')
    # print(file_failed)
    print_list(file_failed)



f = r'C:\Users\Administrator\Desktop\tempt'
# get_emails_from_files(f)

# a_l = find_emails(get_text_clipboard())
# print(find_emails(get_text_clipboard()))
# print(a_l)

url_expr = re.compile(r'')



def eliminate(a_list, b_list):
    a_set = set(a_list)
    b_set = set(b_list)
    return list(a_set-b_set)




def getTYChinaBuyers():

    filename = r'C:\Users\Administrator\Desktop\temp2.txt'
    filename = r'C:\Users\Administrator\Desktop\LEADS-BASE\DEEP-MALA\mala-2022-frchina.txt'
    filename = r'C:\Users\Administrator\Desktop\LEADS-BASE\DEEP-PHI\825fchina-philippine.txt'
    country = "菲律宾"
    special1Reg = re.compile(r'((.*?)\n中国 \| 最新交易时间：.*?\n)(((.*?)\n菲律宾\n\d{1,2}\.\d{1,2}%\n)+)') # teyi purchasers from China page
    special2Reg = re.compile(r'(.*?)\n菲律宾\n\d{1,2}\.\d{1,2}%\n')
    content_list4 =[(tuple[1].title(), [tuple.title() for tuple in special2Reg.findall(tuple[2])], ) for tuple in special1Reg.findall(get_text(filename))] # second is : [tuple for tuple in special2Reg.findall(tuple[2])]

    for i in content_list4:
        print(i)

    result_list = list()
    for i, j in content_list4:
        for buyer in j:
            result_list.append([buyer, i])

    storeList2Excel(result_list, f'buying from China{random.randint(100,999)}.xlsx')

# getTYChinaBuyers()

def get_url_text(url):
    # set headers 
    headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.0.0 Safari/537.36'}
    headers['Host'] = 'movie.douban.com'
    try:
        r = requests.get(url, headers=headers, timeout=10)
        print('Status code: ', r.status_code)
        r.raise_for_status()
        return r.text
    except:
        return ''

def parse_lxml(txt, movie_list):
    soup = BeautifulSoup(txt, 'lxml')
    div_list = soup.find_all('div', class_='hd')

    for each in div_list:
        movie = each.a.span.text.strip()
        movie_list.append(movie)

    # return movie_list

def task1():
    movie_list = []
    url0 = 'https://movie.douban.com/top250?start='
    for i in range(10):
        url = url0 + str(25*(i))
        txt = get_url_text(url)
        parse_lxml(txt, movie_list)

    return movie_list


# txt = get_text_clipboard()
# # print(txt)
# # rr = re.compile(r'(\d){1,2}\.\s(.*?)\s')
# rr = re.compile(r"(.*? )\(.*?\)")
# for i in rr.findall(txt):
#     print(i)

# print(rr.findall(txt))
# print(task1())





def find_QQ_from_clipboard():
    txt = get_text_clipboard()
    # qq = [1-9][0-9]{4,}
    qq = re.compile(r'[1-9][0-9]{8,10}')
    cache_set = set()
    for i in qq.findall(txt):
        if i not in cache_set and i != '672321324' and i != '2748852962':
            print(i)
            cache_set.add(i)

    print(f'\nQQ found: {len(cache_set)}')

    final_lt = list(cache_set)
    final_str = f'QQ found: {len(cache_set)}\n\n' + '\n'.join(final_lt)
    pyperclip.copy(final_str)

# find_QQ_from_clipboard()


t = 'Hi {a}, we are sea shipping forwarder in China, does your company have any shippment to enquiry lately?'
t1 = 'Hi {a}, we are sea shipping forwarder in China, if you have any shipping lately, we can connect, thanks, {a}!'
t2 = '''Hi {a}, I browsed your facebook and I'm just wondering if we could be friends online. Sorry if I disturbed you, {a}, but I think you're pretty and kind-hearted that's why I'd like to see if there's a chance.'''
t3 = '''Hi, we are very interest in your {a_lower}, wondering if we can cooperate, you can add my wechat: no2linear, but currently I failed to login, or one of my Chinese partner's wechat is {b}, you can add him to talk. '''
t4 = '''Hi {a}, we are very interest in your company, wondering if we can cooperate, you can add my wechat: no2linear, but currently I failed to login, or one of my Chinese partner's wechat is {b}, you can add him to talk. And we can connect here first on Linkedin, {a}! '''


# t_dzy = '''Hi {a}, I happened to have 1000 UK e-cigarette clients/leads from the UK with all the email addresses, phone numbers and their websites, that I can sell to you USD 1 or RMB 6.5, if you're interested, you can add my forwarder Bill's wechat: mangocutting, he can sell it to you. Bill is our forwarder that has advantages in Australia and Malaysia.'''
t_dzy = '''Hi {a}!I happen to have 325 vaping customers/prospects from the UK, with all email addresses, phone numbers and their websites, I can sell you for $1 or RMB 10 for all, if you're interested, add my freight forwarder Bill's WeChat: "mangocutting" and he can sell it to you on behalf of me.'''
t_furniture = '''Hi {a}, I happen to have 325 furniture customers/prospects from Australia, with all email addresses, phone numbers and their websites, I can sell you for $1 or RMB 10 for all, if you're interested, add my freight forwarder Bill's WeChat: "mangocutting" and he can sell it to you on behalf of me.'''


t1 = 'Hi {a}, we are sea shipping forwarder in China, wondering, if your company have any shippment to enquiry lately? Thanks for reading my message, {a}!'
t2 = 'Hello {a}, we are sea shipping forwarder in China, if you have any shipping lately, feel free to ask me, Thanks for reading my message, {a}!'
t3 = 'Dear {a}, I just checked your linkedin profile, and I just wonder if your company ever need shipping services? And we happend to be shipping forwarders in China that have advantanges in your country. Thanks for reading my message, {a}!'


t_forwarder = '''Hi {a}, we are very interest in your company, wondering if we can cooperate, you can add my wechat: no2linear, but currently I failed to login, or one of my Chinese shipping forwarder's wechat: anihc_kcuf, you can add him to talk. '''



def c_plus(t_list):
    while True:
        time.sleep(0.2)
        t = random.choice(t_list)

        new_copy = pyperclip.paste()
        random_p = random.randint(1,100)

        b = 'anihc_kcuf'
        # if random_p % 3 == 0:
        #     b = 'notyet54321'
        # elif random_p % 3 == 1:
        #     b = 'anihc_kcuf'
        # else:
        #     b = 'mangocutting'


        if len(new_copy) < 10 and new_copy.strip() != '':
            d = {'a': new_copy.title().strip(), 'b': b, 'a_lower': new_copy.lower().strip()}
            pyperclip.copy(t.format(**d))
        elif new_copy.strip() == '':
            d = {'a': '', 'b': '', 'a_lower': ''}
            pyperclip.copy(t.format(**d))


def f(lower_bound, upper_bound, t=0.3, times=10):
    import time, random
    for i in range(times):
        time.sleep(t)
        r = random.random()
        if r < lower_bound or r > upper_bound:
            break
    if i == times - 1:
        print('Times runs out', i)
        r = 0
    return r - 0.5

def f2(a,b,n, t=0.3, times=10):
    total = 0 
    for i in range(n):
        total += f(a,b, t)
    return total / n


def remove_duplicates(a_str, b_str=''):
    a_set = set(a_str.lower().split())
    b_set = set(a_str.lower().split())
    r_set = a_set - b_set
    r_str = ' '.join(list(r_set))
    print(r_str)
    return r_str




def get_list_fr_file(filename):
    with open(filename) as fo:
        temp_list = fo.readlines()
        result = list(filter(lambda a: len(a)>0, map(lambda a: a.strip().lower(), temp_list)))
    return result


def get_longest_string(alist):
    t_len = 0
    t_str = ''

    for i in alist:
        if len(i) > t_len:
            t_len = len(i)
            t_str = i
    return t_str

print(get_longest_string([]))


def get_names_fr_emails(txt, names_file, default_name='manager', mode=1):


    names_list = get_list_fr_file(names_file)
    emails_list =  find_emails(txt, display=False)
    # print(len(emails_list))
    # print_list(emails_list)
    email_usernames = [e.split('@')[0] for e in emails_list]
    # print(email_usernames)
    names_found = []
    for email_username in email_usernames:
        # checking each email_username 
        temp_list_for_names = []
        for name in names_list:
            if email_username.startswith(name):
                temp_list_for_names.append(name)

        name = get_longest_string(temp_list_for_names).title()
        if not name:
            name = default_name
        print(name)
        names_found.append(name)

    if mode == 1:    
        pyperclip.copy('\n'.join(names_found))
        # print('Copied names.')

    if mode == 1:
        return names_found
    else:
        emails_list



def classify(alist):
    alist.sort(key=lambda a: len(a))
    result_list = []
    while len(alist) > 0:
        temp = alist.pop()
        temp_lt = [temp]
        for i in alist[:]:
            if i in temp:
                temp_lt.append(i)
                alist.remove(i)
        result_list.append(temp_lt)
    return result_list







import pyperclip
import re

def purify_string(text):
    # Remove empty lines
    text = '\n'.join(line.strip() for line in text.splitlines() if line.strip())

    # Reduce redundant spaces
    text = re.sub(r'\s{2,}', ' ', text)

    # Add proper punctuation (assuming each sentence ends with a period)
    text = re.sub(r'(?<=\w)\s+(?=[.!?])', ' ', text)

    return text

def process_clipboard():
    # Get text from the clipboard
    clipboard_text = pyperclip.paste()

    # Purify the string
    purified_text = purify_string(clipboard_text)

    # Copy the purified text back to the clipboard
    pyperclip.copy(purified_text)

if __name__ == "__main__":
    # process_clipboard()
    find_emails(txt=pyperclip.paste())



# if __name__ == '__main__':
#     import os
#     # print()
#     # print(os.getcwd())
#     os.chdir(r'C:\Users\tianx\Downloads\project2022-master\project2022-master')

#     # get_names_fr_emails(get_text_clipboard(), names_file='names.txt',mode=1)
#     find_emails(get_text_clipboard(), display=True)
#     # get_words_frequency(get_text_clipboard())
    # txt = get_text_clipboard()
    # lt = txt.split('\n')
    # ltt = list(filter(lambda x: not ('未' in x), lt))
    # print_list(ltt)



'''





'''