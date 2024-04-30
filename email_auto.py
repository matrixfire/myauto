import openpyxl, re, time, random, sys, json, os, datetime
import smtplib

from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart

from email.mime.application import MIMEApplication 
from email.utils import formataddr
import traceback
import threading

import pyinputplus as pyip
from pathlib import Path
from jinja2 import Template




email_subjects = [
    "Explore New Horizons with Reobrix Building Blocks!",
    "Reobrix Building Blocks: Unleash Your Creativity!",
    "Build Your Dreams with Reobrix Building Blocks!",
    "Reobrix News: Latest Innovations in Building Blocks!",
    "Transform Your World with Reobrix Building Blocks!",
    "Discover the Art of Building with Reobrix Blocks!",
    "Reobrix Building Blocks: Perfect for Every Builder!",
    "Get Creative with Reobrix’s Newest Building Blocks!",
    "Why Reobrix Building Blocks Are Right for You!",
    "Elevate Your Building Experience with Reobrix!"
]



class EmailClient:
    """ SMTP email client.
    Connect and log in to an SMTP server, send an email, and disconnect from the server.
    """
    def __init__(self, smtp_domain, port):
        self.smtp_obj = None
        self.sender = ''
        self.email_counter = 0
        self.smtp_domain = smtp_domain
        self.port = port

    def connect_tls(self):
        """Connect to an SMTP Server using STARTTLS."""
        self.smtp_obj = smtplib.SMTP(self.smtp_domain, self.port)
        self.smtp_obj.ehlo()  # Sending the SMTP “Hello” message
        self.smtp_obj.starttls()  # Starting TLS encryption

    def connect_ssl(self):
        """Connect to an SMTP Server using SSL.
        If the smtplib.SMTP() call is not successful, your SMTP server might not support TLS on port 587.
        In this case, you will need to create an SMTP object using smtplib.SMTP_SSL() and port 465 instead.
        """
        self.smtp_obj = smtplib.SMTP_SSL(self.smtp_domain, self.port)
        self.smtp_obj.ehlo()

    def login(self, email_address, email_password=""):
        """Log in to the SMTP server."""
        if email_password == '':
            email_password = pyip.inputPassword(f'Password for email({email_address}): ')
        try:
            self.smtp_obj.login(email_address, email_password)  # Logging in to the SMTP server
            print(f'{email_address}: logged in')
        except Exception as e:
            print(f'{email_address}, login failed: {str(e)}')

    def send(self, from_address, to_address, msg):
        """Send an email."""
        self.smtp_obj.sendmail(from_address, to_address, msg)
        self.email_counter += 1

    def disconnect(self):
        """Disconnect from the SMTP server."""
        self.smtp_obj.quit()
        print('Email SMTP server quitted!')




text = "Hello,\nThis is a plain text email.\nRegards,\nYour Name"
sent_from = "matrixbox@qq.com"
send_to = "bill@reobrix.com"

# Create MIMEText object
msg = MIMEText(text, 'plain')
msg['Subject'] = 'Simple Plain Text Email'
msg['From'] = sent_from
msg['To'] = send_to


# # Create the MIME multipart message
# msg = MIMEMultipart('alternative')
# msg['Subject'] = 'An interesting email'
# msg['From'] = sent_from
# msg['To'] = ', '.join(send_to)

# # Attach both plain text and HTML parts to the message
# part_plain = MIMEText(text, 'plain')
# part_html = MIMEText(html, 'html')
# msg.attach(part_plain)
# msg.attach(part_html)


# ec = EmailClient()
# ec.connect_ssl('smtp.qq.com', 465)
# ec.login('349500371@qq.com', 'fuutvkptczbrbhcd')
# # ec.send('349500371@qq.com', 'bill@reobrix.com', 'Subject: So long.\nDear Alice, so long and thanks for all the fish.\nSincerely, Bob')
# ec.send('349500371@qq.com', 'bill@reobrix.com', msg.as_string())
# ec.disconnect()



def get_stored_info(json_history_file):
    """
    Retrieves a list containing emails sent, stored in a JSON file.
    """
    filename = os.path.join('temp_email', json_history_file)
    print('SAVE FILE TO', filename)
    try:
        with open(filename) as f:
            emailsSent = json.load(f)
    except FileNotFoundError:
        print(f'Json file: {json_history_file} not found')
        return []
    else:
        if 'yahoo' in json_history_file:
            print('YAHOO EMAIL HISTORY', emailsSent)
        return emailsSent


def save_email_history(email, emailsSent, json_history_file):
    """
    Saves the emails list back to the JSON file.
    """
    current_time = datetime.datetime.now()
    time_str = f'{current_time.year}-{current_time.month}-{current_time.day}'
    filename = os.path.join('temp_email', json_history_file)
    print(f'Email({email}) saved in json file.')
    emailsSent.append((email, time_str))
    with open(filename, 'w') as f:
        json.dump(emailsSent, f)




def extractDataFromExcel(templatefile, sheetNumber=0, header=False):
    """
    Extracts data from an Excel workbook.

    Args:
        templatefile (str): Path to the Excel workbook.
        sheetNumber (int): The sheet number to extract data from, 1-based index.
        header (bool): If True, skips the first row assuming it is the header.

    Returns:
        list: A list of lists containing the extracted data. Each inner list represents a row.

    Raises:
        FileNotFoundError: If the specified file does not exist.
        IndexError: If the sheet number is out of range.
        Exception: For any other issues that arise during execution.
    """
    try:
        # Load the workbook and select the specified sheet by index
        wb = openpyxl.load_workbook(templatefile)
        print(f'Sheet names list: {wb.sheetnames}')
        if sheetNumber:
            sheet = wb[wb.sheetnames[sheetNumber - 1]]
        else:
            sheet = wb.active

        # Extract data from the sheet
        cache_list = []
        for i in range(1, sheet.max_row + 1):  # Iterate through each row
            temp_list_for_row = []
            for j in range(1, sheet.max_column + 1):  # Iterate through each column in the row
                temp_list_for_row.append(sheet.cell(row=i, column=j).value)
            cache_list.append(temp_list_for_row)

        # Skip the first row if header is True
        return cache_list[1:] if not header else cache_list

    except FileNotFoundError:
        print("Error: The file does not exist. Please check the file path.")
        sys.exit(1)
    except IndexError:
        print("Error: Sheet number is out of range. Please check the sheet number.")
        sys.exit(1)
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        sys.exit(1)



def autoEmail(emailClient, email, title, msg, emailsSent, json_history_file):
    """
    Sends an email using specified client settings, appends the sender signature, and logs the history.
    Parameters:
        emailClient (EmailClient): The email client instance used for sending emails.
        email (str): The recipient's email address.
        title (str): The subject line of the email.
        msg (str): The body of the email.
        emailsSent (list): A list of sent emails to be logged.
        json_history_file (str): The filename for storing email log history.
    """
    sender = emailClient.sender
    sender_name = emailClient.sender_name

    if 'notyet54321' not in msg:
        msg += f"\nRegards\n{sender_name}\nKingtrans Container Line(Shenzhen) CO., LTD"

    m = MIMEText(msg, 'plain', 'utf8')
    m['From'] = sender
    m['To'] = email
    m['Subject'] = title

    try:
        sendmailStatus = emailClient.server.sendmail(sender, email, m.as_string())
    except smtplib.SMTPServerDisconnected:
        print(f'{emailClient.sender} caught it: disconnect!!')
        emailClient.connect()
        emailClient.login()
        sendmailStatus = emailClient.send(sender, email, m.as_string())

    if sendmailStatus:
        print(f'{emailClient.sender}: There was a problem sending email to {email}: {sendmailStatus}')
    else:
        print('-' * 100)
        print(f'Title: {title}, {sender} -> {email}')
        print(msg)
        print(f'\n{sender} -> {email}: {title} - ok')
        current_time = datetime.datetime.now()
        print(f"{current_time.strftime('%H:%M:%S')}")
        print('-' * 100)
        save_email_history(email, emailsSent, json_history_file)


    


def getAccountInfo(myAccounts):
    """
    Generate a dictionary of accounts and the number of emails sent today.
    Parameters:
        myAccounts (list): List of accounts to check.
    Returns:
        dict: A dictionary mapping each email account to the number of emails sent today.
    """
    result_dict_raw = {acc[0]: get_stored_info(f"{acc[0].split('@')[0]}-{acc[-1]}.json") for acc in myAccounts}
    return {email: get_num(jsondata) for email, jsondata in result_dict_raw.items()}


def ask(myAccounts, templatefile):
    """
    Interactively ask the user to choose an email account and Excel tab for email processing.
    Parameters:
        myAccounts (list): List of email accounts available.
        templatefile (str): Path to the Excel template file.
    Returns:
        tuple: Selected email account, domain name, Excel tab ID, and messaging mode.
    """
    sheets = openpyxl.load_workbook(templatefile).sheetnames
    result_dict = getAccountInfo(myAccounts)

    for i, acc in enumerate(myAccounts):
        print(f'Enter {i+1} for {acc[0]}, emails already sent today: {result_dict[acc[0]]}')

    multiMsg = pyip.inputNum(prompt='Enter mode (Empty or 1 for single, 2 for multiple, 3-5 for groups): ', blank=True, min=1, max=5)
    emailAcc, emailDomainName = None, None
    if multiMsg not in [3, 4, 5]:
        num = pyip.inputNum('Email account number:', min=1, lessThan=len(myAccounts) + 1)
        emailAcc = myAccounts[num - 1]
        emailDomainName = emailAcc[2]  # Or use split method on emailAcc[0] as needed.

    for i, tab in enumerate(sheets):
        print(f'Enter {i+1} for {tab}')
    excelTabId = pyip.inputNum(f'Enter tab ID 1-{len(sheets)}: ', min=1, lessThan=len(sheets) + 1)

    return emailAcc, emailDomainName, excelTabId, multiMsg




# initial settings

# qq = ['amazingtransition1@qq.com', 'lzeigqinjesxbegg', 'qq']
# qq2 = ['349500371@qq.com', 'fuutvkptczbrbhcd', 'qq']
# qq3 = ['mangocutting@qq.com', 'znczqetqojqidcfa', 'qq']
# qq4 = ['divideandconquer@qq.com', 'mbabhgdhdchdbhhh', 'qq']
# qq5 = ['475853055@qq.com', 'vvwzavwulcvybhib', 'qq']


# # e163 = ['matrixbox@163.com', 'GWNRCXEXINIVGQQA', '163']
# E163 = ['mangocutting@163.com', 'EXXUDFPSVFXHPASH', '163']
# ee1_163 = ['ivoq1690cb@163.com', 'ex535491', '163']
# ee2_163 = ['tgraed451pql@163.com', 'sv620835', '163']


# yahoo = ['matrixbox@yahoo.com', 'vtioleoeammkhhuy', 'yahoo']
# # yahoo2 = ['greatinequality@yahoo.com', 'cvxvttvowldnyazz']

# tencent = ['bill.zou@kingtrans.com.cn', 'Ks2022', 'kingtrans']



# ali = ['bill.zou@szkingtrans.com', 'Alihtam51409', 'szkingtrans']

# tencent2 = ['sales06@cnkingtrans.com', 'Tthtam51409', 'cnkingtrans']
# qy163 = ['sales06@kingtrans.cc', 'kingtrans123', 'qiye163']
# newEmail1 = ['sales06@kingtrans.ltd', 'Elhtam51409', 'kaifang']
# newEmail2 = ['sales09@kingtrans.ltd', 'Elhtam51409', 'kaifang']


# # mix = ['bill@mix.com', 'mix']
# yandex1 = ['whynotbinary@yandex.com', 'lgmflbuqtqqozhkp', 'yandex']
# # yandex2 = ['recursivedream@yandex.com', 'azxffljchocjwjex', 'yandex'] # may spam
# # yandex3 = ['greatiteration@yandex.com', 'rreekalaxwugdqhn', 'yandex']  # may spam


# e126 = ['recursivesolution@126.com', 'GWRQEDXRJEROECVO']
# aol = ['binarysearch01@aol.com', 'kiahlbyxmempunlc']
# sohu = ['binarysearch01@sohu.com', 'L01KRIHMFTO']
# outlook = ['recursivedream@outlook.com', ''] #300 emails zhantingle! no use!
# outlook2 = ['algorithmfuture@outlook.com', 'Elhtam51409'] #300 emails
# gmail = ['mangocutting@gmail.com', 'ynfobiwnrcpnrelj']
# e139 = ['matrixfire@139.com', 'b8e20770c1597ccb2000']



# gmail2 = ['zmeng049@gmail.com', 'BILL0415', 'gmail']

# # zohomail = ['matrixbox@zohomail.com', 'YLN4rvqeZNJx']

# '''ivoq1690cb@163.com----ex535491
# tgraed451pql@163.com----sv620835'''

# myAccounts = []
# myAccounts.append(tencent)
# myAccounts.append(ali)
# myAccounts.append(tencent2)
# myAccounts.append(qq)    
# myAccounts.append(qq2)
# myAccounts.append(qq3)    
# myAccounts.append(qq4)    
# myAccounts.append(qq5)      
# myAccounts.append(yahoo)    
# myAccounts.append(qy163)
# myAccounts.append(newEmail1)
# myAccounts.append(newEmail2)

# myAccounts.append(gmail2)
# groupA = [tencent2, ali, tencent] # 3
# groupB = [qq, qq2, qq3, qq4, qq5] # , yahoo] 4
# groupC = [tencent, newEmail2] #, yandex1, outlook2] 5 qy163, 




def muiltiEmails(infoFromExcel, group):
    """
    Manages sending emails using multiple threads, each handling a part of the list.
    Parameters:
        infoFromExcel (list): A list of email information extracted from Excel.
        group (list): A list of email account groups.
    """
    group_dict = divideList2Parts(infoFromExcel, len(group))
    thread_list = []
    for i, emailAccounts in enumerate(group):
        thisEmailClient = EmailClient(emailAccounts[-1], emailAccounts)
        thread = threading.Thread(target=emailWork, args=(thisEmailClient, group_dict[i]))
        thread_list.append(thread)

    for thread in thread_list:
        time.sleep(5)  # Controlled start delay to avoid sudden surge
        thread.start()




def extract_email(text):
    email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    match = re.search(email_pattern, text)
    return match.group(0) if match else ''



def email_worker(sender_address, sender_password="", smtp_info=None):
    # 1, Initialize global counter and mutex for thread synchronization
    global_counter = 0
    # 2, Define base directory and file paths; in other words, where does my recipients email addresses sit
    base_dir = 'temp_email'
    templatefile_name = 'email_auto.xlsx'
    templatefile = os.path.join(base_dir, templatefile_name)

    # 3, Get the recipients emails along with its associated info
    sheet_num = pyip.inputInt("excel tab: ", blank=True, default=0)
    # if not sheet_num:
    #     sheet_num = 0
    infoFromExcel = extractDataFromExcel(templatefile, sheet_num) # for test reason, I just extract first tab

    recipients_emails_list = [extract_email(" ".join([str(i) for i in sublist])) for sublist in infoFromExcel]
    recipients_emails_list = list(filter(lambda x: len(x) > 0, recipients_emails_list))

    print(recipients_emails_list)


    # 4, decide what mode to use to send bulk emails: the default mode is using one email to send, with random contents;

    # 4.1 , set up the sender account
    # ec = EmailClient('smtp.qq.com', 465)
    ec = EmailClient(smtp_info[0], smtp_info[1])
    ec.connect_ssl()

    if not sender_password:
        sender_password = pyip.inputPassword(f'Password for email({sender_address}): ')
    ec.login(sender_address, sender_password)
    # sender_address = 'matrixbox@qq.com'
    # sender_address = 'amazingtransition1@qq.com'
    # ec.login(sender_address, 'fuutvkptczbrbhcd')
    
    
    # 4.2, prepare the contents
    path = Path('words1.json')
    words_str = path.read_text()
    words = json.loads(words_str)

    path2 = Path(base_dir+'/signature.txt')
    signature_html = path2.read_text(encoding='utf-8')
    # print(signature_html)

    path3 = Path(base_dir+'/email_body.txt')
    body_html = path3.read_text(encoding='utf-8')
    print(body_html)

    # Create MIMEText object
    # 5, bulking sending
    for email_addr in recipients_emails_list[:]:
        subject = random.choice(email_subjects)
        text = random.choice(words)

        # msg = MIMEText(text, 'plain')
        msg = MIMEMultipart('alternative') # html
        msg['Subject'] = subject
        msg['From'] = sender_address
        msg['To'] = email_addr
        msg['Reply-To'] = 'bill@reobrix.com'

        body_html = body_html.replace("{{email_signature}}", signature_html) # html
        template = Template(body_html)
        html_content = template.render(email_body=text)

        part_text = MIMEText(text, 'plain')
        part_html = MIMEText(html_content, 'html') # html


        msg.attach(part_text) # html
        msg.attach(part_html) # html

        # ec.send(sender_address, email_addr, msg.as_string())
        try:
            ec.smtp_obj.send_message(msg)
        except smtplib.SMTPServerDisconnected:
            ec.connect_ssl()
            ec.login(sender_address, sender_password)            
            ec.smtp_obj.send_message(msg)

        time_2_wait = random.randint(21, 300)
        print(f'{sender_address}->{email_addr}\nSubjects: {subject}\n{text}\n')

        global_counter += 1
        print(f'Email(s) already sent: {global_counter}.\nWaiting for {time_2_wait}secs.')
        time.sleep(time_2_wait)

    ec.disconnect()



if __name__ == '__main__':
    # email_worker('amazingtransition1@qq.com')
    # email_worker('mangocutting@163.com', smtp_info=('smtp.163.com', 465))
    email_worker(pyip.inputEmail("Email Address: "), smtp_info=(pyip.inputStr("smtp server: "), pyip.inputInt("smtp serverport(default 465): ", default=465, blank=True)))





# d = {"work":[60,5, 20], "js":[10, 5, 10], "chatgpt": [10, 5, 10], "machine learning basics": [10, 5, 5], "digital marketing learning": [10, 5,10]}
# import random
# random_jobs = []
# for i, lt in d.items():
#     for _ in range(lt[0]):
#         random_jobs.append(i)
# def f():
#     job = random.choice(random_jobs)
#     print(f"{job}: {random.randint(d[job][1], d[job][2])}")



'''
Notes:

import itertools

list1 = [1, 2]
list2 = ['a', 'b']
perms = list(itertools.permutations(list1))
product = list(itertools.product(list1, list2))



'''